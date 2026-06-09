import base64
from datetime import timedelta
from io import BytesIO
import logging
from pathlib import Path
import uuid
from xml.sax.saxutils import escape

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Count, Q
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_POST
from django.utils.timezone import localtime, now
from django.utils.translation import gettext as _
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as ReportLabImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .defaults import (
    DEFAULT_TEACHER_EMAIL,
    DEFAULT_TEACHER_FIRST_NAME,
    DEFAULT_TEACHER_LAST_NAME,
    DEFAULT_TEACHER_PASSWORD,
    DEFAULT_TEACHER_USERNAME,
    LEGACY_DEFAULT_TEACHER_USERNAME,
)
from .forms import AbsenceJustificationForm, ClassScheduleForm, CustomAuthenticationForm, DocumentProcessingForm, DocumentRequestForm, EtudiantForm, FiliereForm, ImportClassesExcelForm, ImportExcelForm, JustificationReviewForm, PasswordChangeForm, SignUpForm, UserProfileUpdateForm
from .models import AbsenceJustification, ClassSchedule, DocumentRequest, Etudiant, Filiere, Notification, Presence, QRAttendanceSession

# Configure logging
logger = logging.getLogger(__name__)

WEEKDAY_ORDER = {
    'Monday': 0,
    'Tuesday': 1,
    'Wednesday': 2,
    'Thursday': 3,
    'Friday': 4,
    'Saturday': 5,
    'Sunday': 6,
}
SUPPORTED_LANGUAGES = {'en', 'fr'}

# Importer openpyxl pour lire les fichiers Excel
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Alternative: utiliser xlrd pour les anciens fichiers Excel
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


# ======================== LANGUAGE DETECTION ========================

def _normalize_language_code(language):
    if not language:
        return None
    normalized = language.split('-')[0].lower()
    return normalized if normalized in SUPPORTED_LANGUAGES else None


def _get_preferred_language(request):
    """Resolve the user's preferred language with a predictable fallback."""
    language = _normalize_language_code(request.session.get('django_language'))
    if language:
        return language

    browser_language = request.META.get('HTTP_ACCEPT_LANGUAGE', '')
    if browser_language:
        language = _normalize_language_code(browser_language.split(',')[0])
        if language:
            return language

    return 'fr'


def detect_language_and_redirect(request, target_url='login'):
    """
    Detect user's preferred language and redirect to language-prefixed URL.
    Checks in order: session, browser Accept-Language, default language
    """
    language = _get_preferred_language(request)

    # If user is logged in and trying to access login, redirect to dashboard
    if target_url == 'login' and request.user.is_authenticated:
        return redirect_after_login(request.user)
    
    # Redirect to language-prefixed URL
    return redirect(f'/{language}/{target_url}/')


def root_language_redirect(request):
    """
    Redirect root URL to language-prefixed home page.
    Detects language from session, browser, or defaults to French.
    """
    if request.user.is_authenticated:
        return redirect_after_login(request.user)
    return detect_language_and_redirect(request, 'login')


@cache_control(public=True, max_age=86400)
def logo_file(request):
    logo_path = get_logo_path()
    return FileResponse(logo_path.open('rb'), content_type='image/jpeg')


# ======================== HELPER FUNCTIONS ========================

def get_student_for_user(user):
    """Safely get student profile for a user."""
    username = getattr(user, 'username', 'anonymous')
    try:
        if not user.is_authenticated:
            return None
        return getattr(user, 'etudiant_profile', None)
    except Exception as e:
        logger.error(f"Error getting student profile for user {username}: {str(e)}")
        return None


def get_portal_students_for_user(user):
    """Return the student records that should feed the student portal."""
    student = get_student_for_user(user)
    if not student:
        return Etudiant.objects.none()

    student_ids = [student.id]
    if student.nom:
        duplicate_ids = list(
            Etudiant.objects.filter(nom=student.nom).values_list('id', flat=True)
        )
        student_ids.extend(duplicate_ids)

    return Etudiant.objects.filter(id__in=set(student_ids)).select_related('filiere', 'user')


def is_admin_user(user):
    return bool(getattr(user, 'is_authenticated', False) and (user.is_staff or user.is_superuser))


def redirect_after_login(user):
    """Redirect user to appropriate page after login."""
    username = getattr(user, 'username', 'anonymous')
    try:
        student = get_student_for_user(user)
        if student:
            return redirect('student_dashboard')
        if is_admin_user(user) and user.notifications.filter(is_read=False).exists():
            return redirect('notifications')
        if is_teacher_user(user):
            return redirect('calendar')
        return redirect('dashboard')
    except Exception as e:
        logger.error(f"Error redirecting user {username}: {str(e)}")
        return redirect('dashboard')


def validate_file_size(file, max_size_mb=5):
    """Validate uploaded file size."""
    if file.size > max_size_mb * 1024 * 1024:
        return False, _("File is too large. Maximum size is {max_size_mb}MB").format(max_size_mb=max_size_mb)
    return True, _("File size is valid")


def validate_excel_file(file):
    """Validate Excel file structure."""
    try:
        if not file.name.lower().endswith(('.xlsx', '.xls')):
            return False, _("Invalid file format. Please upload .xlsx or .xls file")
        return True, _("File format is valid")
    except Exception as e:
        logger.error(f"Error validating Excel file: {str(e)}")
        return False, _("Error validating file")


def get_attendance_stats(filiere):
    """Get attendance statistics for a filiere."""
    try:
        today = now().date()
        total_students = filiere.etudiants.count()
        attendance_counts = Presence.objects.filter(
            etudiant__filiere=filiere,
            date=today,
        ).aggregate(
            present_count=Count('id', filter=Q(present=True)),
            absent_count=Count('id', filter=Q(present=False)),
        )
        present_today = attendance_counts['present_count']
        absent_today = attendance_counts['absent_count']
        
        attendance_rate = 0
        if total_students > 0:
            attendance_rate = (present_today / total_students) * 100
            
        return {
            'total': total_students,
            'present': present_today,
            'absent': absent_today,
            'rate': round(attendance_rate, 2)
        }
    except Exception as e:
        logger.error(f"Error getting attendance stats for filiere {filiere.id}: {str(e)}")
        return {'total': 0, 'present': 0, 'absent': 0, 'rate': 0}


def annotate_filiere_attendance(queryset, target_date):
    return queryset.annotate(
        student_count=Count('etudiants', distinct=True),
        present_count=Count(
            'etudiants__presences',
            distinct=True,
            filter=Q(
                etudiants__presences__date=target_date,
                etudiants__presences__present=True,
            ),
        ),
        absent_count=Count(
            'etudiants__presences',
            distinct=True,
            filter=Q(
                etudiants__presences__date=target_date,
                etudiants__presences__present=False,
            ),
        ),
    )


def build_monthly_student_totals(students, current_date):
    monthly_counts = Presence.objects.filter(
        etudiant__in=students,
        date__month=current_date.month,
        date__year=current_date.year,
    ).values('etudiant_id').annotate(
        total_presences=Count('id', filter=Q(present=True)),
        total_absences=Count('id', filter=Q(present=False)),
    )
    return {
        row['etudiant_id']: {
            'total_presences': row['total_presences'],
            'total_absences': row['total_absences'],
        }
        for row in monthly_counts
    }


def log_user_action(user, action, details=""):
    """Log user actions for audit trail."""
    try:
        log_message = f"User {user.username} - Action: {action}"
        if details:
            log_message += f" - Details: {details}"
        logger.info(log_message)
    except Exception as e:
        logger.error(f"Error logging user action: {str(e)}")


def split_full_name(full_name):
    """Split a full name into first and last name parts for Django's user model."""
    parts = str(full_name or '').strip().split()
    if not parts:
        return '', ''
    if len(parts) == 1:
        return parts[0], ''
    return parts[0], ' '.join(parts[1:])


def is_teacher_user(user):
    return bool(
        getattr(user, 'is_authenticated', False)
        and not get_student_for_user(user)
        and not is_admin_user(user)
    )


def get_teacher_identifiers(user):
    identifiers = {
        (getattr(user, 'username', '') or '').strip(),
        (user.get_full_name() or '').strip(),
        f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip(),
        (user.first_name or '').strip(),
    }
    return {identifier for identifier in identifiers if identifier}


def get_teacher_schedule_queryset(user):
    base_queryset = ClassSchedule.objects.select_related('filiere')
    if is_admin_user(user):
        return base_queryset
    if not is_teacher_user(user):
        return base_queryset.none()

    assigned_class_ids = list(
        Filiere.objects.filter(teacher=user).values_list('id', flat=True)
    )
    if assigned_class_ids:
        return base_queryset.filter(filiere_id__in=assigned_class_ids)

    filters = Q()
    for identifier in get_teacher_identifiers(user):
        filters |= Q(professor__iexact=identifier)

    if not filters:
        return base_queryset.none()
    return base_queryset.filter(filters)


def get_teacher_filieres_queryset(user):
    if is_admin_user(user):
        return Filiere.objects.all()
    if not is_teacher_user(user):
        return Filiere.objects.none()

    assigned_filieres = Filiere.objects.filter(teacher=user)
    assigned_filiere_ids = list(assigned_filieres.values_list('id', flat=True))
    if assigned_filiere_ids:
        return assigned_filieres

    return Filiere.objects.filter(
        id__in=get_teacher_schedule_queryset(user).values_list('filiere_id', flat=True)
    ).distinct()


def user_can_access_filiere(user, filiere):
    if is_admin_user(user):
        return True
    if not is_teacher_user(user):
        return False
    return get_teacher_filieres_queryset(user).filter(id=filiere.id).exists()


def get_roster_source_filieres(exclude_filiere=None):
    queryset = Filiere.objects.annotate(student_total=Count('etudiants')).filter(student_total__gt=0)
    if exclude_filiere:
        queryset = queryset.exclude(id=exclude_filiere.id)
    return queryset.order_by('nom')


def normalize_excel_student_row(row, row_idx):
    nom = row[0] if len(row) > 0 else None
    numero_etudiant = row[1] if len(row) > 1 else None
    email = row[2] if len(row) > 2 else None

    if not nom:
        return None, _("Ligne {row_idx}: Le nom de l'etudiant est obligatoire").format(row_idx=row_idx)

    nom = str(nom).strip()
    numero_etudiant = str(numero_etudiant).strip() if numero_etudiant else None
    email = str(email).strip() if email else ''

    if not nom or nom.lower() == 'none':
        return None, _("Ligne {row_idx}: Le nom de l'etudiant ne peut pas etre vide").format(row_idx=row_idx)

    if numero_etudiant and numero_etudiant.lower() == 'none':
        numero_etudiant = None
    if email and email.lower() == 'none':
        email = ''

    return {
        'nom': nom,
        'numero_etudiant': numero_etudiant,
        'email': email,
    }, None


def upsert_student_from_row(filiere, row_data):
    if row_data['numero_etudiant']:
        _, created = Etudiant.objects.update_or_create(
            filiere=filiere,
            numero_etudiant=row_data['numero_etudiant'],
            defaults={
                'nom': row_data['nom'],
                'email': row_data['email'],
            },
        )
        return created

    student, created = Etudiant.objects.get_or_create(
        filiere=filiere,
        nom=row_data['nom'],
        defaults={'email': row_data['email']},
    )
    if not created and row_data['email'] and student.email != row_data['email']:
        student.email = row_data['email']
        student.save(update_fields=['email'])
    return created


def import_students_from_worksheet(filiere, worksheet):
    imported_count = 0
    updated_count = 0
    errors = []

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_data, error = normalize_excel_student_row(row, row_idx)
            if error:
                errors.append(error)
                continue

            if upsert_student_from_row(filiere, row_data):
                imported_count += 1
            else:
                updated_count += 1
        except Exception as e:
            errors.append(_("Ligne {row_idx}: {error}").format(row_idx=row_idx, error=str(e)))

    return imported_count, updated_count, errors


def copy_students_between_classes(source_filiere, target_filiere):
    source_students = list(source_filiere.etudiants.order_by('nom'))
    existing_numbers = set(
        target_filiere.etudiants.exclude(numero_etudiant__isnull=True)
        .exclude(numero_etudiant='')
        .values_list('numero_etudiant', flat=True)
    )
    existing_name_emails = set(
        target_filiere.etudiants.values_list('nom', 'email')
    )

    students_to_create = []
    skipped_count = 0

    for source_student in source_students:
        if source_student.numero_etudiant:
            if source_student.numero_etudiant in existing_numbers:
                skipped_count += 1
                continue
            existing_numbers.add(source_student.numero_etudiant)
        else:
            key = (source_student.nom, source_student.email)
            if key in existing_name_emails:
                skipped_count += 1
                continue
            existing_name_emails.add(key)

        students_to_create.append(
            Etudiant(
                nom=source_student.nom,
                email=source_student.email,
                numero_etudiant=source_student.numero_etudiant,
                profile_photo=source_student.profile_photo,
                filiere=target_filiere,
            )
        )

    Etudiant.objects.bulk_create(students_to_create)
    return len(students_to_create), skipped_count


def ensure_default_teacher_account():
    """
    Keep a predictable teacher demo account available for the app.
    """
    should_reset_password = getattr(settings, 'RESET_DEFAULT_TEACHER_PASSWORD', False)
    legacy_user = User.objects.filter(username=LEGACY_DEFAULT_TEACHER_USERNAME).first()
    if legacy_user and not User.objects.filter(username=DEFAULT_TEACHER_USERNAME).exists():
        legacy_user.username = DEFAULT_TEACHER_USERNAME
        legacy_user.first_name = DEFAULT_TEACHER_FIRST_NAME
        legacy_user.last_name = DEFAULT_TEACHER_LAST_NAME
        legacy_user.email = DEFAULT_TEACHER_EMAIL
        legacy_user.is_staff = False
        legacy_user.is_superuser = False
        legacy_user.set_password(DEFAULT_TEACHER_PASSWORD)
        legacy_user.save()
        return legacy_user

    teacher_user, created = User.objects.get_or_create(
        username=DEFAULT_TEACHER_USERNAME,
        defaults={
            'email': DEFAULT_TEACHER_EMAIL,
            'first_name': DEFAULT_TEACHER_FIRST_NAME,
            'last_name': DEFAULT_TEACHER_LAST_NAME,
        },
    )

    fields_to_update = []
    if created or should_reset_password:
        teacher_user.set_password(DEFAULT_TEACHER_PASSWORD)
        fields_to_update.append('password')
    if teacher_user.is_staff:
        teacher_user.is_staff = False
        fields_to_update.append('is_staff')
    if teacher_user.is_superuser:
        teacher_user.is_superuser = False
        fields_to_update.append('is_superuser')
    if not teacher_user.email:
        teacher_user.email = DEFAULT_TEACHER_EMAIL
        fields_to_update.append('email')
    if teacher_user.first_name != DEFAULT_TEACHER_FIRST_NAME:
        teacher_user.first_name = DEFAULT_TEACHER_FIRST_NAME
        fields_to_update.append('first_name')
    if teacher_user.last_name != DEFAULT_TEACHER_LAST_NAME:
        teacher_user.last_name = DEFAULT_TEACHER_LAST_NAME
        fields_to_update.append('last_name')

    if created:
        teacher_user.save()
    elif fields_to_update:
        teacher_user.save(update_fields=fields_to_update)

    return teacher_user


def create_notification(user, title, message, level=Notification.LEVEL_INFO, link=''):
    if not user:
        return
    Notification.objects.create(
        user=user,
        title=title,
        message=message,
        level=level,
        link=link,
    )


def create_admin_notifications(title, message, level=Notification.LEVEL_INFO, link=''):
    admin_users = User.objects.filter(Q(is_staff=True) | Q(is_superuser=True)).distinct()
    notifications = [
        Notification(
            user=admin_user,
            title=title,
            message=message,
            level=level,
            link=link,
        )
        for admin_user in admin_users
    ]
    if notifications:
        Notification.objects.bulk_create(notifications)


def send_absence_email(student, date, filiere):
    """Send an absence alert email to the student. Silently skipped if no email configured."""
    email_address = (student.email or '').strip()
    if not email_address:
        return
    threshold = getattr(settings, 'ABSENCE_ALERT_THRESHOLD', 3)
    if threshold > 0 and student.total_absences < threshold:
        return
    subject = _("Absence enregistree - {classe}").format(classe=filiere.nom)
    body = _(
        "Bonjour {nom},\n\n"
        "Vous avez ete marque(e) absent(e) le {date} pour la classe {classe}.\n"
        "Total d'absences : {total}\n\n"
        "Si vous avez une justification, connectez-vous a AttendX pour la soumettre.\n\n"
        "-- AttendX"
    ).format(
        nom=student.nom,
        date=date.strftime('%d/%m/%Y'),
        classe=filiere.nom,
        total=student.total_absences,
    )
    try:
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [email_address], fail_silently=True)
    except Exception as exc:
        logger.warning("Absence email failed for %s: %s", email_address, exc)


def get_logo_path():
    return Path(settings.BASE_DIR) / 'LOGO2.jpeg'


def build_dashboard_analytics(today):
    absences_today = Presence.objects.filter(date=today, present=False).select_related('etudiant', 'etudiant__filiere')
    classes = list(annotate_filiere_attendance(Filiere.objects.all(), today).order_by('nom'))
    top_absent_students = (
        Etudiant.objects.order_by('-total_absences', 'nom').select_related('filiere')[:5]
    )
    class_breakdown = [
        {
            'filiere': filiere,
            'present': filiere.present_count,
            'absent': filiere.absent_count,
            'rate': round((filiere.present_count / filiere.student_count) * 100) if filiere.student_count else 0,
        }
        for filiere in classes
    ]

    return {
        'absences_today': absences_today[:8],
        'top_absent_students': top_absent_students,
        'class_breakdown': class_breakdown[:6],
        'needs_attention_count': absences_today.count(),
    }


def build_student_analytics(presences):
    monthly = {}
    for presence in presences:
        key = presence.date.strftime('%m/%Y')
        monthly.setdefault(key, {'present': 0, 'absent': 0})
        monthly[key]['present' if presence.present else 'absent'] += 1

    trend = []
    for label, totals in list(monthly.items())[:6]:
        total = totals['present'] + totals['absent']
        trend.append({
            'label': label,
            'present': totals['present'],
            'absent': totals['absent'],
            'rate': round((totals['present'] / total) * 100) if total else 0,
        })
    return trend


def build_schedule_section(filiere, schedules, current_day_name, current_time):
    """Build an ordered weekly calendar summary for a filiere."""
    current_day_rank = WEEKDAY_ORDER.get(current_day_name, 99)
    assigned_teacher_name = ''
    if getattr(filiere, 'teacher', None):
        assigned_teacher_name = filiere.teacher.get_full_name() or filiere.teacher.username
    ordered_schedules = sorted(
        schedules,
        key=lambda schedule: (
            WEEKDAY_ORDER.get(schedule.day_of_week, 99),
            schedule.start_time,
            schedule.end_time,
        ),
    )

    grouped_days = []
    day_lookup = {}
    next_session = None
    today_session_count = 0

    for schedule in ordered_schedules:
        schedule.display_professor = schedule.professor or assigned_teacher_name
        schedule.display_date = schedule.class_date.strftime('%d/%m/%Y') if schedule.class_date else ''
        if schedule.day_of_week == current_day_name:
            today_session_count += 1

        if next_session is None:
            schedule_day_rank = WEEKDAY_ORDER.get(schedule.day_of_week, 99)
            is_future_day = schedule_day_rank > current_day_rank
            is_later_today = schedule.day_of_week == current_day_name and schedule.start_time >= current_time
            if is_future_day or is_later_today:
                next_session = schedule

        label = schedule.get_day_of_week_display()
        if label not in day_lookup:
            day_lookup[label] = {
                'label': label,
                'count': 0,
                'items': [],
            }
            grouped_days.append(day_lookup[label])
        day_lookup[label]['items'].append(schedule)
        day_lookup[label]['count'] += 1

    first_session = ordered_schedules[0] if ordered_schedules else None
    last_session = ordered_schedules[-1] if ordered_schedules else None
    if next_session is None and ordered_schedules:
        next_session = first_session

    return {
        'filiere': filiere,
        'grouped_days': grouped_days,
        'total_sessions': len(ordered_schedules),
        'occupied_days': len(grouped_days),
        'today_session_count': today_session_count,
        'first_session': first_session,
        'last_session': last_session,
        'next_session': next_session,
    }


# ======================== VIEWS ========================

def home(request):
    """Home page - redirect to appropriate dashboard."""
    if request.user.is_authenticated:
        return redirect_after_login(request.user)
    return redirect('login')


def login_view(request):
    ensure_default_teacher_account()

    if request.user.is_authenticated:
        return redirect_after_login(request.user)

    form = CustomAuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        login(request, user)
        return redirect_after_login(user)

    return render(
        request,
        'login.html',
        {
            'form': form,
            'default_teacher_credentials': {
                'username': DEFAULT_TEACHER_USERNAME,
                'password': DEFAULT_TEACHER_PASSWORD,
            },
        },
    )


def signup_view(request):
    if request.user.is_authenticated:
        return redirect_after_login(request.user)

    form = SignUpForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(request, _("Votre compte etudiant a ete cree avec succes."))
        return redirect('student_dashboard')

    return render(request, 'signup.html', {'form': form})


@login_required
def logout_view(request):
    log_user_action(request.user, "LOGOUT")
    logout(request)
    return redirect('login')


@login_required
def user_profile(request):
    """Display user profile page."""
    try:
        if is_teacher_user(request.user):
            messages.error(request, _("Le compte enseignant n'a acces qu'au planning et a l'appel."))
            return redirect('calendar')
        context = {
            'user': request.user,
            'student': get_student_for_user(request.user),
        }
        log_user_action(request.user, "VIEW_PROFILE")
        return render(request, 'user_profile.html', context)
    except Exception as e:
        logger.error(f"Error loading user profile for {request.user.username}: {str(e)}")
        messages.error(request, _("Error loading profile page"))
        return redirect('dashboard')


@login_required
def change_password(request):
    """Change user password."""
    try:
        if is_teacher_user(request.user):
            messages.error(request, _("Le compte enseignant n'a acces qu'au planning et a l'appel."))
            return redirect('calendar')
        student = get_student_for_user(request.user)
        if request.method == 'POST':
            form = PasswordChangeForm(request.user, request.POST)
            if form.is_valid():
                user = request.user
                user.set_password(form.cleaned_data['new_password1'])
                user.save()
                messages.success(request, _("Votre mot de passe a été changé avec succès."))
                log_user_action(user, "CHANGE_PASSWORD")
                logout(request)
                return redirect('login')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            form = PasswordChangeForm(request.user)
        
        return render(request, 'change_password.html', {'form': form})
    except Exception as e:
        logger.error(f"Error changing password for user {request.user.username}: {str(e)}")
        messages.error(request, _("Erreur lors du changement de mot de passe"))
        return redirect('dashboard')


@login_required
def update_profile(request):
    """Update user profile information."""
    try:
        if is_teacher_user(request.user):
            messages.error(request, _("Le compte enseignant n'a acces qu'au planning et a l'appel."))
            return redirect('calendar')
        student = get_student_for_user(request.user)
        if request.method == 'POST':
            form = UserProfileUpdateForm(request.POST, request.FILES, instance=request.user)
            uploaded_photo = request.FILES.get('profile_photo')
            photo_saved = False

            if uploaded_photo:
                if not student:
                    messages.error(request, _("Aucun profil etudiant n'est lie a ce compte."))
                    return redirect('update_profile')

                extension = Path(uploaded_photo.name).suffix.lower()
                allowed_extensions = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
                allowed_types = {'image/jpeg', 'image/png', 'image/webp', 'image/gif'}
                content_type = getattr(uploaded_photo, 'content_type', '')

                if content_type not in allowed_types and extension not in allowed_extensions:
                    messages.error(request, _("Choisissez une image JPG, PNG, WEBP ou GIF."))
                    return render(request, 'update_profile.html', {'form': form, 'student': student})

                if uploaded_photo.size > UserProfileUpdateForm.MAX_PHOTO_SIZE:
                    messages.error(request, _("La photo ne doit pas depasser 10 MB."))
                    return render(request, 'update_profile.html', {'form': form, 'student': student})

                student.profile_photo = uploaded_photo
                student.save(update_fields=['profile_photo'])
                photo_saved = True
            if form.is_valid():
                form.save()
                messages.success(request, _("Votre profil a été mis à jour avec succès."))
                log_user_action(request.user, "UPDATE_PROFILE")
                return redirect('user_profile')
            if photo_saved:
                messages.success(request, _("Votre photo de profil a ete mise a jour."))
                log_user_action(request.user, "UPDATE_PROFILE_PHOTO")
                return redirect('user_profile')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            form = UserProfileUpdateForm(instance=request.user)
        
        return render(request, 'update_profile.html', {'form': form, 'student': student})
    except Exception as e:
        logger.error(f"Error updating profile for user {request.user.username}: {str(e)}")
        messages.error(request, _("Erreur lors de la mise à jour du profil"))
        return redirect('dashboard')


@login_required
def dashboard(request):
    if get_student_for_user(request.user):
        return redirect('student_dashboard')
    if is_teacher_user(request.user):
        return redirect('calendar')

    today = now().date()
    filieres = annotate_filiere_attendance(Filiere.objects.all(), today).select_related('teacher').order_by('nom')
    filiere_form = FiliereForm(request.POST or None)

    if request.method == 'POST':
        if filiere_form.is_valid():
            filiere = filiere_form.save()
            class_date = filiere_form.cleaned_data.get('class_date')
            day_of_week = filiere_form.cleaned_data.get('day_of_week')
            start_time = filiere_form.cleaned_data.get('start_time')
            end_time = filiere_form.cleaned_data.get('end_time')
            subject = (filiere_form.cleaned_data.get('subject') or '').strip()
            session_type = filiere_form.cleaned_data.get('session_type') or ClassSchedule.SESSION_NORMAL
            room = (filiere_form.cleaned_data.get('room') or '').strip()

            if subject and start_time and end_time and (class_date or day_of_week):
                teacher_name = ''
                if filiere.teacher:
                    teacher_name = filiere.teacher.get_full_name() or filiere.teacher.username
                ClassSchedule.objects.create(
                    filiere=filiere,
                    class_date=class_date,
                    day_of_week=day_of_week,
                    start_time=start_time,
                    end_time=end_time,
                    subject=subject,
                    session_type=session_type,
                    room=room,
                    professor=teacher_name,
                )
            messages.success(request, _("La classe a ete ajoutee avec succes."))
            return redirect('dashboard')
        messages.error(request, _("Impossible d'ajouter la classe. Verifiez le formulaire."))

    stats = {
        'filiere_count': filieres.count(),
        'student_count': Etudiant.objects.count(),
        'attendance_count_today': Presence.objects.filter(date=today).count(),
        'pending_justifications_count': AbsenceJustification.objects.filter(status=AbsenceJustification.STATUS_PENDING).count(),
        'document_requests_count': DocumentRequest.objects.filter(status__in=[DocumentRequest.STATUS_PENDING, DocumentRequest.STATUS_PROCESSING]).count(),
    }
    analytics = build_dashboard_analytics(today)
    saved_filieres = filieres
    return render(
        request,
        'home.html',
        {
            'filieres': filieres,
            'saved_filieres': saved_filieres,
            'filiere_form': filiere_form,
            'stats': stats,
            'today': today,
            'analytics': analytics,
        },
    )


@login_required
def import_classes_excel(request):
    """Importe une classe depuis le nom du fichier Excel et y ajoute les etudiants."""
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')
    if not is_admin_user(request.user):
        messages.error(request, _("Acces reserve a l'administration."))
        return redirect('calendar')

    form = ImportClassesExcelForm(request.POST or None, request.FILES or None)
    saved_classes = Filiere.objects.annotate(student_total=Count('etudiants')).order_by('nom')

    if request.method == 'POST' and form.is_valid():
        if not OPENPYXL_AVAILABLE:
            messages.error(request, _("openpyxl n'est pas installe. Installez-le avec: pip install openpyxl"))
            return render(request, 'import_classes.html', {'form': form, 'saved_classes': saved_classes})

        excel_file = form.cleaned_data['excel_file']
        class_name = Path(excel_file.name).stem.strip()

        try:
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active

            if not class_name:
                messages.error(request, _("Le nom du fichier ne peut pas etre vide."))
                return render(request, 'import_classes.html', {'form': form, 'saved_classes': saved_classes})

            filiere, class_created = Filiere.objects.get_or_create(
                nom=class_name,
                defaults={
                    'description': _("Classe creee depuis le fichier {filename}.").format(filename=excel_file.name),
                },
            )

            imported_count, updated_count, errors = import_students_from_worksheet(filiere, worksheet)

            if class_created:
                message_text = _("Classe {class_name} creee depuis le nom du fichier. ").format(class_name=class_name)
            else:
                message_text = _("Classe {class_name} mise a jour depuis le nom du fichier. ").format(class_name=class_name)

            message_text += _(
                "{imported_count} etudiant(s) ajoute(s), {updated_count} mise(s) a jour."
            ).format(imported_count=imported_count, updated_count=updated_count)

            if errors:
                message_text += " " + _("{error_count} erreur(s) rencontree(s).").format(error_count=len(errors))
                for error in errors[:5]:
                    messages.warning(request, error)

            messages.success(request, message_text)
            return redirect('class_detail', filiere_id=filiere.id)

        except Exception as e:
            messages.error(request, _("Erreur lors de la lecture du fichier: {error}").format(error=str(e)))

    return render(request, 'import_classes.html', {'form': form, 'saved_classes': saved_classes})


@login_required
def select_filiere(request):
    if get_student_for_user(request.user):
        return redirect('student_dashboard')

    today = now().date()
    filieres = list(
        annotate_filiere_attendance(get_teacher_filieres_queryset(request.user), today).select_related('teacher').order_by('nom')
    )

    for filiere in filieres:
        filiere.attendance_rate = round((filiere.present_count / filiere.student_count) * 100) if filiere.student_count else 0

    return render(
        request,
        'select_filiere.html',
        {
            'filieres': filieres,
            'today': today,
        },
    )


@login_required
def class_detail(request, filiere_id):
    filiere = get_object_or_404(Filiere, id=filiere_id)
    if get_student_for_user(request.user):
        return redirect('student_dashboard')
    if not user_can_access_filiere(request.user, filiere):
        messages.error(request, _("Vous ne pouvez consulter que vos propres classes."))
        return redirect('select_filiere')

    teacher_mode = is_teacher_user(request.user)
    students = list(filiere.etudiants.all().order_by('nom'))
    student_form = None if teacher_mode else EtudiantForm(request.POST or None, request.FILES or None)
    schedule_form = None if teacher_mode else ClassScheduleForm(request.POST or None)
    source_filieres = get_roster_source_filieres(filiere) if not teacher_mode else Filiere.objects.none()
    today = now().date()
    attendance_map = {
        presence.etudiant_id: presence.present
        for presence in Presence.objects.filter(etudiant__filiere=filiere, date=today)
    }
    present_count = sum(1 for is_present in attendance_map.values() if is_present)
    absent_count = sum(1 for is_present in attendance_map.values() if not is_present)
    student_count = len(students)
    attendance_rate = round((present_count / student_count) * 100) if student_count else 0

    if request.method == 'POST' and request.POST.get('action') == 'add_student':
        if teacher_mode:
            messages.error(request, _("Les enseignants ne peuvent pas modifier la liste des etudiants."))
            return redirect('class_detail', filiere_id=filiere.id)
        if student_form.is_valid():
            student = student_form.save(commit=False)
            student.filiere = filiere
            if student_form.cleaned_data.get('create_student_account'):
                first_name, last_name = split_full_name(student.nom)
                student_user = User.objects.create_user(
                    username=student_form.cleaned_data['student_username'],
                    password=student_form.cleaned_data['student_password'],
                    email=student.email,
                    first_name=first_name,
                    last_name=last_name,
                )
                student.user = student_user
            student.save()
            messages.success(request, _("L'etudiant a ete ajoute a la classe."))
            return redirect('class_detail', filiere_id=filiere.id)
        messages.error(request, _("Le formulaire etudiant contient des erreurs."))

    if request.method == 'POST' and request.POST.get('action') == 'copy_students_from_class':
        if teacher_mode:
            messages.error(request, _("Les enseignants ne peuvent pas modifier la liste des etudiants."))
            return redirect('class_detail', filiere_id=filiere.id)

        source_filiere = get_object_or_404(Filiere, id=request.POST.get('source_filiere'))
        copied_count, skipped_count = copy_students_between_classes(source_filiere, filiere)

        messages.success(
            request,
            _("Copie terminee: {copied_count} etudiant(s) ajoute(s), {skipped_count} deja present(s).").format(
                copied_count=copied_count,
                skipped_count=skipped_count,
            ),
        )
        return redirect('class_detail', filiere_id=filiere.id)

    if request.method == 'POST' and request.POST.get('action') == 'add_schedule':
        if teacher_mode:
            messages.error(request, _("Les enseignants ne peuvent pas modifier le planning de la classe."))
            return redirect('class_detail', filiere_id=filiere.id)
        if schedule_form.is_valid():
            schedule = schedule_form.save(commit=False)
            schedule.filiere = filiere
            schedule.professor = filiere.teacher.get_full_name() or filiere.teacher.username if filiere.teacher else ''
            schedule.save()
            messages.success(request, _("Le planning de la classe a ete mis a jour."))
            return redirect('class_detail', filiere_id=filiere.id)
        messages.error(request, _("Le formulaire de planning contient des erreurs."))

    justifications = AbsenceJustification.objects.filter(
        presence__etudiant__filiere=filiere
    ).select_related('presence', 'presence__etudiant')[:8]
    schedules = filiere.schedules.all().order_by('class_date', 'day_of_week', 'start_time')

    return render(
        request,
        'list_students.html',
        {
            'filiere': filiere,
            'students': students,
            'student_form': student_form,
            'schedule_form': schedule_form,
            'source_filieres': source_filieres,
            'schedules': schedules,
            'teacher_mode': teacher_mode,
            'attendance_map': attendance_map,
            'present_count': present_count,
            'absent_count': absent_count,
            'attendance_rate': attendance_rate,
            'today': today,
            'justifications': justifications,
        },
    )


@login_required
def import_students_excel(request, filiere_id):
    """Importe une liste d'etudiants depuis un fichier Excel"""
    filiere = get_object_or_404(Filiere, id=filiere_id)
    
    # Verifier que ce n'est pas un etudiant qui essaie d'acceder
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')
    if not is_admin_user(request.user):
        messages.error(request, _("Acces reserve a l'administration."))
        return redirect('class_detail', filiere_id=filiere.id)
    
    form = ImportExcelForm(request.POST or None, request.FILES or None)
    source_filieres = get_roster_source_filieres(filiere)
    
    if request.method == 'POST' and form.is_valid():
        if not OPENPYXL_AVAILABLE:
            messages.error(request, _("openpyxl n'est pas installe. Installez-le avec: pip install openpyxl"))
            return render(request, 'import_students.html', {'form': form, 'filiere': filiere, 'source_filieres': source_filieres})
        
        excel_file = form.cleaned_data['excel_file']
        
        try:
            # Charger le fichier Excel
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active
            
            imported_count, updated_count, errors = import_students_from_worksheet(filiere, worksheet)
            
            # Afficher les resultats
            message_text = _(
                "Import termine: {imported_count} etudiant(s) ajoute(s), {updated_count} mise(s) a jour."
            ).format(imported_count=imported_count, updated_count=updated_count)
            if errors:
                message_text += " " + _("{error_count} erreur(s) rencontree(s).").format(error_count=len(errors))
                for error in errors[:5]:  # Afficher les 5 premieres erreurs
                    messages.warning(request, error)
            
            messages.success(request, message_text)
            return redirect('class_detail', filiere_id=filiere.id)
        
        except Exception as e:
            messages.error(request, _("Erreur lors de la lecture du fichier: {error}").format(error=str(e)))
    
    return render(
        request,
        'import_students.html',
        {
            'form': form,
            'filiere': filiere,
            'source_filieres': source_filieres,
        },
    )


@login_required
def mark_attendance(request, filiere_id):
    return redirect('class_detail', filiere_id=filiere_id)


def _update_student_totals(student):
    student.refresh_attendance_totals()


@login_required
@require_POST
def save_attendance(request, filiere_id):
    filiere = get_object_or_404(Filiere, id=filiere_id)
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')
    if not user_can_access_filiere(request.user, filiere):
        messages.error(request, _("Vous ne pouvez marquer l'absence que pour vos propres classes."))
        return redirect('select_filiere')

    students = filiere.etudiants.all()
    today = now().date()

    with transaction.atomic():
        for student in students:
            is_present = f'present_{student.id}' in request.POST
            presence, created = Presence.objects.get_or_create(
                etudiant=student,
                date=today,
                defaults={'present': is_present},
            )
            became_absent = created and not is_present

            if not created and presence.present != is_present:
                became_absent = presence.present and not is_present
                presence.present = is_present
                presence.save(update_fields=['present'])

            _update_student_totals(student)

            if not is_present and student.user:
                create_notification(
                    student.user,
                    _("Absence enregistree"),
                    _("Vous avez ete marque absent le {date} pour la classe {classe}.").format(
                        date=today.strftime('%d/%m/%Y'),
                        classe=student.filiere.nom,
                    ),
                    level=Notification.LEVEL_WARNING,
                    link=reverse('student_dashboard'),
                )
            if not is_present:
                send_absence_email(student, today, filiere)

            if became_absent:
                marker_name = request.user.get_full_name() or request.user.username
                create_admin_notifications(
                    _("Nouvelle absence marquee"),
                    _("{student} a ete marque absent en {classe} le {date} par {teacher}.").format(
                        student=student.nom,
                        classe=student.filiere.nom,
                        date=today.strftime('%d/%m/%Y'),
                        teacher=marker_name,
                    ),
                    level=Notification.LEVEL_WARNING,
                    link=reverse('absences_classes'),
                )

    messages.success(request, _("L'appel du jour a ete enregistre."))
    return generate_daily_pdf(filiere)


@login_required
def monthly_report(request):
    if get_student_for_user(request.user):
        return redirect('student_dashboard')
    if not is_admin_user(request.user):
        messages.error(request, _("Acces reserve a l'administration."))
        return redirect('calendar')

    filieres = Filiere.objects.all().order_by('nom')
    report = []
    current_date = now().date()

    for filiere in filieres:
        students = list(filiere.etudiants.all().order_by('nom'))
        monthly_totals = build_monthly_student_totals(students, current_date)
        data = [
            {
                'etudiant': etudiant,
                'total_presences': monthly_totals.get(etudiant.id, {}).get('total_presences', 0),
                'total_absences': monthly_totals.get(etudiant.id, {}).get('total_absences', 0),
            }
            for etudiant in students
        ]
        report.append({'filiere': filiere, 'data': data})

    return render(request, 'monthly_report.html', {'report': report, 'today': current_date})


PDF_COLORS = {
    'navy': colors.HexColor('#111827'),
    'blue': colors.HexColor('#E5E7EB'),
    'soft_blue': colors.HexColor('#F3F4F6'),
    'line': colors.HexColor('#9CA3AF'),
    'text': colors.HexColor('#111827'),
    'muted': colors.HexColor('#374151'),
    'row': colors.HexColor('#F9FAFB'),
}


def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        'title': ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=PDF_COLORS['navy'],
            spaceAfter=2,
        ),
        'subtitle': ParagraphStyle(
            'ReportSubtitle',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=PDF_COLORS['muted'],
        ),
        'badge': ParagraphStyle(
            'ReportBadge',
            parent=styles['BodyText'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=12,
            alignment=1,
            textColor=PDF_COLORS['text'],
        ),
        'metric_label': ParagraphStyle(
            'MetricLabel',
            parent=styles['BodyText'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=12,
            textColor=PDF_COLORS['muted'],
        ),
        'metric_value': ParagraphStyle(
            'MetricValue',
            parent=styles['BodyText'],
            fontName='Helvetica-Bold',
            fontSize=13,
            leading=16,
            textColor=PDF_COLORS['text'],
        ),
        'section': ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=16,
            textColor=PDF_COLORS['navy'],
            spaceAfter=8,
        ),
        'table_header': ParagraphStyle(
            'TableHeader',
            parent=styles['BodyText'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=PDF_COLORS['text'],
        ),
        'table_cell': ParagraphStyle(
            'TableCell',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=8.5,
            leading=11,
            textColor=PDF_COLORS['text'],
        ),
        'table_cell_center': ParagraphStyle(
            'TableCellCenter',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=8.5,
            leading=11,
            alignment=1,
            textColor=PDF_COLORS['text'],
        ),
        'empty': ParagraphStyle(
            'EmptyState',
            parent=styles['BodyText'],
            fontName='Helvetica-Oblique',
            fontSize=10,
            leading=14,
            textColor=PDF_COLORS['muted'],
        ),
    }


def _pdf_paragraph(value, style):
    return Paragraph(escape(str(value or '-')), style)


def _build_summary_table(summary_items, styles, available_width):
    cells = []
    cell_width = available_width / max(len(summary_items), 1)
    for label, value in summary_items:
        cells.append(
            Table(
                [[
                    Paragraph(escape(str(label)).upper(), styles['metric_label']),
                ], [
                    Paragraph(escape(str(value)), styles['metric_value']),
                ]],
                colWidths=[cell_width],
                style=TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), PDF_COLORS['soft_blue']),
                    ('BOX', (0, 0), (-1, -1), 0.5, PDF_COLORS['line']),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
                    ('TOPPADDING', (0, 1), (-1, 1), 0),
                    ('BOTTOMPADDING', (0, 1), (-1, 1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]),
            )
        )

    return Table(
        [cells],
        colWidths=[cell_width] * len(cells),
        style=TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]),
    )


def _build_report_pdf(filename, title, subtitle, badge, summary_items, table_headers, table_rows, column_widths, empty_message):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
    )
    styles = _pdf_styles()
    story = []
    logo_path = get_logo_path()

    if logo_path.exists():
        logo = ReportLabImage(str(logo_path), width=18 * mm, height=18 * mm)
    else:
        logo = Paragraph('GC', styles['title'])

    header = Table(
        [[
            logo,
            [Paragraph(escape(title), styles['title']), Paragraph(escape(subtitle), styles['subtitle'])],
            Paragraph(escape(badge), styles['badge']),
        ]],
        colWidths=[22 * mm, doc.width - 55 * mm, 33 * mm],
        style=TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (2, 0), (2, 0), PDF_COLORS['soft_blue']),
            ('BOX', (2, 0), (2, 0), 0.5, PDF_COLORS['line']),
            ('ALIGN', (2, 0), (2, 0), 'CENTER'),
            ('LEFTPADDING', (2, 0), (2, 0), 8),
            ('RIGHTPADDING', (2, 0), (2, 0), 8),
            ('TOPPADDING', (2, 0), (2, 0), 6),
            ('BOTTOMPADDING', (2, 0), (2, 0), 6),
            ('LINEBELOW', (0, 0), (-1, 0), 1, PDF_COLORS['line']),
        ]),
    )
    story.append(header)
    story.append(Spacer(1, 6 * mm))
    story.append(_build_summary_table(summary_items, styles, doc.width))
    story.append(Spacer(1, 7 * mm))
    story.append(Paragraph(escape(_("Details de presence")), styles['section']))

    table_data = [[Paragraph(escape(header), styles['table_header']) for header in table_headers]]
    if table_rows:
        for row in table_rows:
            rendered_row = []
            for idx, cell in enumerate(row):
                style = styles['table_cell_center'] if idx > 0 else styles['table_cell']
                rendered_row.append(_pdf_paragraph(cell, style))
            table_data.append(rendered_row)
    else:
        table_data.append([Paragraph(escape(empty_message), styles['empty'])] + [''] * (len(table_headers) - 1))

    table = Table(table_data, colWidths=column_widths, repeatRows=1)
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['blue']),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['line']),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, PDF_COLORS['row']]),
    ]
    if not table_rows:
        table_style.extend([
            ('SPAN', (0, 1), (-1, 1)),
            ('ALIGN', (0, 1), (-1, 1), 'LEFT'),
        ])
    table.setStyle(TableStyle(table_style))
    story.append(table)

    def decorate_page(pdf_canvas, pdf_doc):
        page_width, page_height = A4
        pdf_canvas.saveState()
        pdf_canvas.setFillColor(PDF_COLORS['text'])
        pdf_canvas.rect(0, page_height - 3 * mm, page_width, 3 * mm, fill=1, stroke=0)
        pdf_canvas.setStrokeColor(PDF_COLORS['line'])
        pdf_canvas.line(doc.leftMargin, 12 * mm, page_width - doc.rightMargin, 12 * mm)
        pdf_canvas.setFillColor(PDF_COLORS['muted'])
        pdf_canvas.setFont('Helvetica', 8)
        pdf_canvas.drawString(doc.leftMargin, 8 * mm, title)
        pdf_canvas.drawRightString(page_width - doc.rightMargin, 8 * mm, _("Page {page}").format(page=pdf_doc.page))
        pdf_canvas.restoreState()

    doc.build(story, onFirstPage=decorate_page, onLaterPages=decorate_page)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def generate_monthly_pdf(request, filiere_id):
    filiere = get_object_or_404(Filiere, id=filiere_id)
    if get_student_for_user(request.user):
        return redirect('student_dashboard')
    if not is_admin_user(request.user):
        messages.error(request, _("Acces reserve a l'administration."))
        return redirect('calendar')

    current_date = now().date()
    students = list(filiere.etudiants.all().order_by('nom'))
    monthly_totals = build_monthly_student_totals(students, current_date)
    table_rows = []
    total_presences = 0
    total_absences = 0

    for etudiant in students:
        student_totals = monthly_totals.get(etudiant.id, {})
        student_presences = student_totals.get('total_presences', 0)
        student_absences = student_totals.get('total_absences', 0)
        total_presences += student_presences
        total_absences += student_absences
        total_sessions = student_presences + student_absences
        rate = f"{round((student_presences / total_sessions) * 100)}%" if total_sessions else "-"
        table_rows.append([
            etudiant.nom,
            student_presences,
            student_absences,
            rate,
        ])

    summary_items = [
        (_("Classe"), filiere.nom),
        (_("Etudiants"), len(students)),
        (_("Presences / Absences"), f"{total_presences} / {total_absences}"),
    ]

    return _build_report_pdf(
        filename=f"rapport_{filiere.nom}.pdf",
        title=_("Rapport mensuel - {nom}").format(nom=filiere.nom),
        subtitle=_("Periode : {date}").format(date=current_date.strftime('%m/%Y')),
        badge=_("Mensuel"),
        summary_items=summary_items,
        table_headers=[_("Etudiant"), _("Presences"), _("Absences"), _("Taux")],
        table_rows=table_rows,
        column_widths=[96 * mm, 26 * mm, 26 * mm, 26 * mm],
        empty_message=_("Aucun etudiant disponible pour cette classe."),
    )


def generate_daily_pdf(filiere):
    current_date = now().date()
    students = list(filiere.etudiants.all().order_by('nom'))
    presence_map = {
        presence.etudiant_id: presence.present
        for presence in Presence.objects.filter(etudiant__in=students, date=current_date)
    }
    table_rows = []
    present_count = 0
    absent_count = 0

    for student in students:
        is_present = bool(presence_map.get(student.id))
        status = _("Present") if is_present else _("Absent")
        if is_present:
            present_count += 1
        else:
            absent_count += 1
        table_rows.append([
            student.nom,
            student.numero_etudiant or '-',
            status,
        ])

    summary_items = [
        (_("Classe"), filiere.nom),
        (_("Date"), current_date.strftime('%d/%m/%Y')),
        (_("Present(s) / Absent(s)"), f"{present_count} / {absent_count}"),
    ]

    return _build_report_pdf(
        filename=f"appel_{filiere.nom}_{current_date}.pdf",
        title=_("Appel du jour - {nom}").format(nom=filiere.nom),
        subtitle=_("Synthese de presence du {date}").format(date=current_date.strftime('%d/%m/%Y')),
        badge=_("Journalier"),
        summary_items=summary_items,
        table_headers=[_("Etudiant"), _("Numero"), _("Presence")],
        table_rows=table_rows,
        column_widths=[104 * mm, 38 * mm, 32 * mm],
        empty_message=_("Aucun etudiant disponible pour cette classe."),
    )


@login_required
def student_dashboard(request):
    student = get_student_for_user(request.user)
    if not student:
        return redirect('dashboard')

    portal_students = list(get_portal_students_for_user(request.user))
    presences = Presence.objects.filter(
        etudiant__in=portal_students
    ).select_related('justification', 'etudiant', 'etudiant__filiere').order_by('-date', 'etudiant__nom')
    presence_counts = presences.aggregate(
        present_count=Count('id', filter=Q(present=True)),
        absent_count=Count('id', filter=Q(present=False)),
        justified_count=Count('id', filter=Q(present=False, justification__isnull=False)),
    )
    justified_count = presence_counts['justified_count']
    total_presences = presence_counts['present_count']
    total_absences = presence_counts['absent_count']

    # Get class schedules for all linked student records
    schedule = ClassSchedule.objects.filter(
        filiere__in=[portal_student.filiere for portal_student in portal_students]
    ).order_by('day_of_week', 'start_time', 'filiere__nom').distinct()
    class_names = sorted({portal_student.filiere.nom for portal_student in portal_students})

    stats = {
        'class_name': ', '.join(class_names),
        'presence_count': total_presences,
        'absence_count': total_absences,
        'justified_count': justified_count,
        'pending_absence_count': max(total_absences - justified_count, 0),
    }
    analytics = build_student_analytics(list(presences))
    current_time = localtime(now())
    return render(
        request,
        'student_dashboard.html',
        {
            'student': student,
            'portal_students': portal_students,
            'presences': presences,
            'stats': stats,
            'schedule': schedule,
            'analytics': analytics,
            'current_time': current_time,
        },
    )


@login_required
def justify_absence(request, presence_id):
    student = get_student_for_user(request.user)
    if not student:
        return redirect('dashboard')

    portal_students = list(get_portal_students_for_user(request.user))
    presence = get_object_or_404(Presence, id=presence_id, etudiant__in=portal_students, present=False)
    justification = getattr(presence, 'justification', None)
    form = AbsenceJustificationForm(request.POST or None, request.FILES or None, instance=justification)

    if request.method == 'POST' and form.is_valid():
        justification = form.save(commit=False)
        justification.presence = presence
        justification.status = AbsenceJustification.STATUS_PENDING
        justification.save()
        create_notification(
            request.user,
            _("Justification envoyee"),
            _("Votre justification pour l'absence du {date} a ete envoyee.").format(
                date=presence.date.strftime('%d/%m/%Y')
            ),
            level=Notification.LEVEL_INFO,
            link=reverse('student_dashboard'),
        )
        messages.success(request, _("Votre justification a bien ete envoyee."))
        return redirect('student_dashboard')

    return render(
        request,
        'justify_absence.html',
        {
            'student': student,
            'presence': presence,
            'form': form,
        },
    )


@login_required
def manage_justifications(request):
    """
    Permet aux professeurs de gerer les justifications d'absence (approuver/rejeter).
    """
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')
    if not is_admin_user(request.user):
        messages.error(request, _("Seul l'administrateur peut approuver ou refuser les justifications."))
        return redirect('dashboard')

    pending_justifications = AbsenceJustification.objects.filter(
        status=AbsenceJustification.STATUS_PENDING
    ).select_related('presence__etudiant', 'presence__etudiant__filiere').order_by('-submitted_at')

    if request.method == 'POST':
        form = JustificationReviewForm(request.POST)
        if form.is_valid():
            justification = get_object_or_404(AbsenceJustification, id=form.cleaned_data['justification_id'])
            action = request.POST.get('action') or form.cleaned_data['action']
            teacher_comment = form.cleaned_data['teacher_comment']

            action_map = {
                JustificationReviewForm.ACTION_APPROVE: (
                    AbsenceJustification.STATUS_APPROVED,
                    messages.success,
                    _("Justification de {nom} approuvee."),
                ),
                JustificationReviewForm.ACTION_REJECT: (
                    AbsenceJustification.STATUS_REJECTED,
                    messages.warning,
                    _("Justification de {nom} rejetee."),
                ),
            }
            status, message_level, message_text = action_map[action]
            justification.review(status=status, reviewed_by=request.user, teacher_comment=teacher_comment)
            message_level(request, message_text.format(nom=justification.presence.etudiant.nom))
            if justification.presence.etudiant.user:
                create_notification(
                    justification.presence.etudiant.user,
                    _("Justification mise a jour"),
                    _("Votre justification du {date} a ete {status}.").format(
                        date=justification.presence.date.strftime('%d/%m/%Y'),
                        status=justification.get_status_display().lower(),
                    ),
                    level=Notification.LEVEL_SUCCESS if action == JustificationReviewForm.ACTION_APPROVE else Notification.LEVEL_WARNING,
                    link=reverse('student_dashboard'),
                )
            return redirect('manage_justifications')
        messages.error(request, _("Impossible de traiter la justification."))

    review_forms = {
        justification.id: {
            'approve': JustificationReviewForm(
                initial={'justification_id': justification.id, 'action': JustificationReviewForm.ACTION_APPROVE, 'teacher_comment': justification.teacher_comment}
            ),
            'reject': JustificationReviewForm(
                initial={'justification_id': justification.id, 'action': JustificationReviewForm.ACTION_REJECT, 'teacher_comment': justification.teacher_comment}
            ),
        }
        for justification in pending_justifications
    }

    return render(request, 'manage_justifications.html', {
        'pending_justifications': pending_justifications,
        'review_forms': review_forms,
    })


# ======================== NEW SIDEBAR VIEWS ========================

@login_required
def general_information(request):
    if is_teacher_user(request.user):
        messages.error(request, _("Cette section n'est pas disponible pour le compte enseignant."))
        return redirect('calendar')
    return render(request, 'general_information.html')

@login_required
def calendar_view(request):
    student = get_student_for_user(request.user)
    current_dt = localtime()
    current_day_name = current_dt.strftime('%A')
    current_time = current_dt.time()

    if student:
        filieres = [student.filiere]
        schedule_map = {student.filiere.id: list(student.filiere.schedules.all())}
    elif is_teacher_user(request.user):
        teacher_schedule_queryset = get_teacher_schedule_queryset(request.user).order_by(
            'filiere__nom', 'day_of_week', 'start_time'
        )
        filieres = list(get_teacher_filieres_queryset(request.user).prefetch_related('schedules').order_by('nom'))
        schedule_map = {
            filiere.id: [schedule for schedule in teacher_schedule_queryset if schedule.filiere_id == filiere.id]
            for filiere in filieres
        }
    else:
        filieres = list(Filiere.objects.prefetch_related('schedules').order_by('nom'))
        schedule_map = {filiere.id: list(filiere.schedules.all()) for filiere in filieres}

    calendar_sections = []
    for filiere in filieres:
        schedules = schedule_map.get(filiere.id, [])
        calendar_sections.append(
            build_schedule_section(filiere, schedules, current_day_name, current_time)
        )

    has_any_schedule = any(section['total_sessions'] for section in calendar_sections)

    return render(
        request,
        'calendar.html',
        {
            'student': student,
            'calendar_sections': calendar_sections,
            'has_any_schedule': has_any_schedule,
        },
    )

@login_required
def submitted_documents(request):
    student = get_student_for_user(request.user)
    if not student:
        return redirect('dashboard')

    form = DocumentRequestForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        document_request = form.save(commit=False)
        document_request.etudiant = student
        document_request.save()
        create_notification(
            request.user,
            _("Demande envoyee"),
            _("Votre demande de document \"{doc}\" a ete envoyee.").format(doc=document_request.document_type),
            level=Notification.LEVEL_INFO,
            link=reverse('submitted_documents'),
        )
        messages.success(request, _("Votre demande de document a ete envoyee."))
        return redirect('submitted_documents')

    documents = student.document_requests.all()
    return render(request, 'submitted_documents.html', {'form': form, 'documents': documents, 'student': student})


@login_required
def manage_documents(request):
    if get_student_for_user(request.user):
        return redirect('student_dashboard')
    if not is_admin_user(request.user):
        messages.error(request, _("Acces reserve a l'administration."))
        return redirect('calendar')

    if request.method == 'POST':
        form = DocumentProcessingForm(request.POST, request.FILES)
        if form.is_valid():
            document_request = get_object_or_404(DocumentRequest, id=form.cleaned_data['request_id'])
            action_to_status = {
                DocumentProcessingForm.ACTION_PROCESS: DocumentRequest.STATUS_PROCESSING,
                DocumentProcessingForm.ACTION_READY: DocumentRequest.STATUS_READY,
                DocumentProcessingForm.ACTION_REJECT: DocumentRequest.STATUS_REJECTED,
            }
            action = request.POST.get('action') or form.cleaned_data['action']
            document_request.update_status(
                status=action_to_status[action],
                admin_comment=form.cleaned_data['admin_comment'],
                delivered_file=form.cleaned_data.get('delivered_file'),
            )
            if document_request.etudiant.user:
                create_notification(
                    document_request.etudiant.user,
                    _("Document mis a jour"),
                    _("Votre demande \"{doc}\" est maintenant : {status}.").format(
                        doc=document_request.document_type,
                        status=document_request.get_status_display(),
                    ),
                    level=Notification.LEVEL_SUCCESS if action == DocumentProcessingForm.ACTION_READY else Notification.LEVEL_INFO,
                    link=reverse('submitted_documents'),
                )
            messages.success(request, _("Le statut du document a ete mis a jour."))
            return redirect('manage_documents')
        messages.error(request, _("Impossible de mettre a jour le document."))

    pending_docs = DocumentRequest.objects.select_related('etudiant', 'etudiant__filiere').filter(
        status__in=[DocumentRequest.STATUS_PENDING, DocumentRequest.STATUS_PROCESSING]
    )
    stats = {
        'pending': DocumentRequest.objects.filter(status=DocumentRequest.STATUS_PENDING).count(),
        'processing': DocumentRequest.objects.filter(status=DocumentRequest.STATUS_PROCESSING).count(),
        'ready': DocumentRequest.objects.filter(status=DocumentRequest.STATUS_READY).count(),
    }
    processing_forms = {
        doc.id: {
            'process': DocumentProcessingForm(initial={'request_id': doc.id, 'action': DocumentProcessingForm.ACTION_PROCESS, 'admin_comment': doc.admin_comment}),
            'ready': DocumentProcessingForm(initial={'request_id': doc.id, 'action': DocumentProcessingForm.ACTION_READY, 'admin_comment': doc.admin_comment}),
            'reject': DocumentProcessingForm(initial={'request_id': doc.id, 'action': DocumentProcessingForm.ACTION_REJECT, 'admin_comment': doc.admin_comment}),
        }
        for doc in pending_docs
    }
    return render(request, 'manage_documents.html', {'pending_docs': pending_docs, 'stats': stats, 'processing_forms': processing_forms})

@login_required
def academic_tracking(request):
    if not is_admin_user(request.user):
        messages.error(request, _("Cette section est reservee aux administrateurs."))
        return redirect('dashboard' if is_teacher_user(request.user) else 'student_dashboard')
    return render(request, 'academic_tracking.html')

@login_required
def academic_grades(request):
    if is_teacher_user(request.user):
        messages.error(request, _("Cette section n'est pas disponible pour le compte enseignant."))
        return redirect('calendar')
    return render(request, 'academic_grades.html')

@login_required
def absences_classes(request):
    student = get_student_for_user(request.user)
    if not student:
        if is_teacher_user(request.user):
            messages.error(request, _("Cette section n'est pas disponible pour le compte enseignant."))
            return redirect('calendar')

        filieres = list(
            Filiere.objects.annotate(
                student_count=Count('etudiants', distinct=True),
                absence_count=Count(
                    'etudiants__presences',
                    distinct=True,
                    filter=Q(etudiants__presences__present=False),
                ),
                justified_absence_count=Count(
                    'etudiants__presences',
                    distinct=True,
                    filter=Q(
                        etudiants__presences__present=False,
                        etudiants__presences__justification__isnull=False,
                    ),
                ),
            ).order_by('nom')
        )
        absences = (
            Presence.objects.filter(present=False)
            .select_related('etudiant', 'etudiant__filiere', 'justification')
            .order_by('-date', 'etudiant__filiere__nom', 'etudiant__nom')
        )
        absences_by_class = {}
        for absence in absences:
            absences_by_class.setdefault(absence.etudiant.filiere_id, []).append(absence)

        for filiere in filieres:
            filiere.absence_records = absences_by_class.get(filiere.id, [])
            filiere.unjustified_absence_count = max(
                filiere.absence_count - filiere.justified_absence_count,
                0,
            )

        stats = {
            'total': sum(filiere.absence_count for filiere in filieres),
            'justified': sum(filiere.justified_absence_count for filiere in filieres),
            'unjustified': sum(filiere.unjustified_absence_count for filiere in filieres),
            'classes': len(filieres),
        }
        return render(
            request,
            'absences_classes.html',
            {
                'admin_mode': True,
                'filieres': filieres,
                'stats': stats,
            },
        )

    portal_students = list(get_portal_students_for_user(request.user))
    presences = Presence.objects.filter(
        etudiant__in=portal_students,
        present=False,
    ).select_related('justification', 'etudiant', 'etudiant__filiere').order_by('-date')
    absence_counts = presences.aggregate(
        total_count=Count('id'),
        justified_count=Count('id', filter=Q(justification__isnull=False)),
    )
    total_absences = absence_counts['total_count']
    justified_count = absence_counts['justified_count']
    stats = {
        'total': total_absences,
        'justified': justified_count,
        'unjustified': max(total_absences - justified_count, 0),
    }
    return render(request, 'absences_classes.html', {'student': student, 'presences': presences, 'stats': stats})

@login_required
def absences_tests(request):
    return render(request, 'absences_tests.html')

@login_required
def absences_exams(request):
    return render(request, 'absences_exams.html')

@login_required
def cheating(request):
    return render(request, 'cheating.html')

@login_required
def sent_sms(request):
    if not is_admin_user(request.user):
        messages.error(request, _("Acces reserve a l'administration."))
        return redirect('calendar')
    notifications = Notification.objects.filter(level=Notification.LEVEL_WARNING).select_related('user')[:25]
    return render(request, 'sent_sms.html', {'notifications': notifications})


@login_required
def notifications_view(request):
    if is_teacher_user(request.user):
        messages.error(request, _("Le compte enseignant n'a acces qu'au planning et a l'appel."))
        return redirect('calendar')
    notifications = request.user.notifications.all()
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return render(request, 'notifications.html', {'notifications': notifications})


# ═══════════════════════════════════════════════════════════════════
#  FEATURE 1 — Analytics JSON endpoint  (consumed by Chart.js)
# ═══════════════════════════════════════════════════════════════════

@login_required
def analytics_data(request):
    """Return JSON data for the dashboard charts."""
    if get_student_for_user(request.user):
        return JsonResponse({'error': 'forbidden'}, status=403)

    today = now().date()

    # ── 6-month attendance trend ──────────────────────────────────
    trend = []
    for months_ago in range(5, -1, -1):
        first = (today.replace(day=1) - timedelta(days=months_ago * 28)).replace(day=1)
        counts = Presence.objects.filter(
            date__year=first.year,
            date__month=first.month,
        ).aggregate(
            present_count=Count('id', filter=Q(present=True)),
            absent_count=Count('id', filter=Q(present=False)),
        )
        trend.append({
            'label': first.strftime('%m/%Y'),
            'present': counts['present_count'] or 0,
            'absent': counts['absent_count'] or 0,
        })

    # ── Per-class attendance rate (all-time) ──────────────────────
    filieres = Filiere.objects.annotate(
        total_records=Count('etudiants__presences', distinct=True),
        present_records=Count(
            'etudiants__presences',
            distinct=True,
            filter=Q(etudiants__presences__present=True),
        ),
    ).order_by('nom')

    class_rates = [
        {
            'label': f.nom,
            'rate': round(f.present_records / f.total_records * 100) if f.total_records else 0,
            'total': f.total_records,
        }
        for f in filieres
    ]

    # ── Weekly absence heatmap (last 4 weeks) ─────────────────────
    weekly = []
    for weeks_ago in range(3, -1, -1):
        week_start = today - timedelta(days=today.weekday() + 7 * weeks_ago)
        week_end = week_start + timedelta(days=6)
        absent_count = Presence.objects.filter(
            date__range=(week_start, week_end),
            present=False,
        ).count()
        weekly.append({
            'label': week_start.strftime('%d/%m'),
            'absences': absent_count,
        })

    return JsonResponse({
        'trend': trend,
        'class_rates': class_rates,
        'weekly': weekly,
    })


# ═══════════════════════════════════════════════════════════════════
#  FEATURE 3 — QR Code Attendance
# ═══════════════════════════════════════════════════════════════════

@login_required
def qr_session(request, filiere_id):
    """Teacher view: generate / display a QR code for the class."""
    try:
        import qrcode as qrcode_lib
    except ImportError:
        messages.error(request, _("Le module qrcode n'est pas installe. Lancez: pip install qrcode[pil]"))
        return redirect('class_detail', filiere_id=filiere_id)

    filiere = get_object_or_404(Filiere, id=filiere_id)
    if get_student_for_user(request.user):
        messages.error(request, _("Acces non autorise."))
        return redirect('student_dashboard')
    if not user_can_access_filiere(request.user, filiere):
        messages.error(request, _("Vous ne pouvez acceder qu'a vos propres classes."))
        return redirect('select_filiere')

    today = now().date()
    session_duration = timedelta(minutes=15)

    if request.method == 'POST' and request.POST.get('action') == 'refresh':
        # Teacher explicitly asked for a new token
        QRAttendanceSession.objects.filter(filiere=filiere, date=today).delete()

    session, created = QRAttendanceSession.objects.get_or_create(
        filiere=filiere,
        date=today,
        defaults={
            'created_by': request.user,
            'expires_at': now() + session_duration,
        },
    )

    # Renew if expired
    if not session.is_valid():
        session.token = uuid.uuid4()
        session.expires_at = now() + session_duration
        session.save(update_fields=['token', 'expires_at'])

    scan_url = request.build_absolute_uri(
        reverse('qr_scan', kwargs={'token': str(session.token)})
    )

    # Build QR image as base64 PNG
    qr = qrcode_lib.QRCode(
        version=None,
        error_correction=qrcode_lib.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(scan_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#073b88', back_color='white')
    buf = BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    scanned_today = Presence.objects.filter(
        etudiant__filiere=filiere,
        date=today,
        present=True,
    ).select_related('etudiant').order_by('etudiant__nom')

    return render(request, 'qr_attendance.html', {
        'filiere': filiere,
        'session': session,
        'qr_b64': qr_b64,
        'scan_url': scan_url,
        'scanned_today': scanned_today,
        'today': today,
    })


@login_required
def qr_scan(request, token):
    """Student view: mark themselves present by visiting the QR link."""
    session = get_object_or_404(QRAttendanceSession, token=token)
    student = get_student_for_user(request.user)

    if not student:
        messages.error(request, _("Seuls les etudiants peuvent utiliser ce lien de presence."))
        return redirect('dashboard')

    if not session.is_valid():
        return render(request, 'qr_scan.html', {
            'expired': True,
            'session': session,
            'student': student,
        })

    if student.filiere_id != session.filiere_id:
        return render(request, 'qr_scan.html', {
            'wrong_class': True,
            'session': session,
            'student': student,
            'expected_class': session.filiere.nom,
        })

    today = session.date
    presence, created = Presence.objects.get_or_create(
        etudiant=student,
        date=today,
        defaults={'present': True},
    )
    if not created and not presence.present:
        presence.present = True
        presence.save(update_fields=['present'])

    student.refresh_attendance_totals()

    return render(request, 'qr_scan.html', {
        'success': True,
        'already_marked': not created and presence.present,
        'session': session,
        'student': student,
        'today': today,
    })
